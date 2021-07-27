import openpyxl


def matrix_production(dict_matrix):
    new_matrix = []
    for value in dict_matrix.values():
        new_matrix.append(value)
    return new_matrix


def transposed(matrix):
    count = 0
    new_matrix = []
    while True:
        new_string = []
        for string in matrix:
            if (count+1) > len(string):
                new_string.append('0')
            else:
                new_string.append(string[count])
        if new_string == ['0','0','0','0']:
            break
        else:
            new_matrix.append(list(new_string))
            count += 1
    return new_matrix


def data_to_write(key:str, value:list):
    pre_matrix = matrix_production(value)
    pre_matrix_1 = transposed(pre_matrix)
    for string in pre_matrix_1:
        string.insert(0, key)
    return pre_matrix_1


def filling_stor_loc(work_sheet)->dict:
    stor_loc = {}
    for row in range(3, work_sheet.max_row + 1):
        name = work_sheet[row][1].value
        if name == None:
            name = 'None'
        quantity = work_sheet[row][2].value
        stor_bin = work_sheet[row][3].value
        print(name, quantity, stor_bin)
        if name == 'None' and quantity==stor_bin==None: # Условие, которое прекращает цикл если все ячейки строчки пустые
            break
        if stor_bin not in stor_loc:
            stor_loc[stor_bin] = {name: [quantity]}
        else:
            dict_s = stor_loc[stor_bin]
            if name in dict_s:
                list_s = dict_s[name]
                list_s.append(quantity)
            else:
                dict_s[name] = [quantity]
    return stor_loc


def unpacking_xlsx(name:str, name_sheet:str):
    name = name + '.xlsx'
    work_book = openpyxl.open(name, read_only=True)
    work_sheet = work_book[name_sheet]
    return work_sheet


def filling_check_dict(check_dict:dict, input_dict:dict, number:str)->dict:
    name_dict = 'stor_loc_' + number
    for sell, content in input_dict.items():
        if sell in check_dict:
            temp_dict = check_dict[sell]
            temp_dict[name_dict] = content
        else:
            check_dict[sell] = {name_dict: content}
    return check_dict


def check_2(input_dict): # функция получает sell_content на вход и выдает 3 множества, general_set в котором записана общая номенклатура и name_1_rock_set и name_2_rock_set в которых хранится номенклатура не пересекающаяся между двумя таблицами
    general_set = set() # множество содержит номенклатуру, которая есть в двух таблицах
    name_1_rock_set = set() # множество содержит номенклатуру, которая есть в первой таблице
    name_tabs = list(input_dict.keys())
    name_1 = name_tabs[0]
    name_2 = name_tabs[1]
    dict_name_1 = input_dict[name_1]
    dict_name_2 = input_dict[name_2]
    for name_rock in dict_name_1:
        if name_rock in dict_name_2:
            general_set.add(name_rock)
        else:
            name_1_rock_set.add(name_rock)
    name_2_rock_set = set(dict_name_2.keys()) - general_set
    return_dict = {'general_set': general_set, 'name_1_rock_set': name_1_rock_set, 'name_2_rock_set': name_2_rock_set}
    return return_dict # возвращает словарь, где ключ название множества, а значение само множество


def check_3(input_dict:dict, general_set:set)->dict:
    name_tabs = list(input_dict.keys())
    name_tab_1 = name_tabs[0]
    name_tab_2 = name_tabs[1]
    tab_1_dict = input_dict[name_tab_1]
    tab_2_dict = input_dict[name_tab_2]
    amount_general = []
    difference_rock_list = []
    for rock in general_set:
        amount_rock_1 = sum(tab_1_dict[rock])
        amount_rock_2 = sum(tab_2_dict[rock])
        if abs(amount_rock_1-amount_rock_2) <=5: # добавлен диапазон разницы в весе(5кг)
            amount_general.append(rock)
        else:
            dif_amn_rock = abs(amount_rock_1 - amount_rock_2)
            difference_rock_list.append(str(rock + ': ' + str(dif_amn_rock)))
    return_dict = {'amount_general': amount_general, 'difference_rock_list': difference_rock_list}
    return return_dict


sheet_book_1 = unpacking_xlsx('test_1', 'Лист1')
sheet_book_2 = unpacking_xlsx('test_2', 'Лист1')
stor_loc_1 = filling_stor_loc(sheet_book_1) # структура данных, которая содержит данные первой таблицы
stor_loc_2 = filling_stor_loc(sheet_book_2) # структура данных, которая содержит данные второй таблицы
# полученные словари имеют структуру {ячейка: {номенклатура: [список с массами]}}


new_dict = {}
s_control = filling_check_dict(check_dict=new_dict ,input_dict=stor_loc_1, number='1')
new_control_dict = filling_check_dict(s_control, stor_loc_2, '2')

# создаем новую книгу excel
new_book = openpyxl.Workbook()
# обращаемся у странице книги
sheet_new_book = new_book.active
# создаем шапку для новой таблицы
sheet_new_book.cell(row=1,column=1).value = 'cell'
sheet_new_book.cell(row=1,column=2).value = 'matching_data'
sheet_new_book.cell(row=1,column=3).value = 'mismatched_data'
sheet_new_book.cell(row=1,column=4).value = 'first_table_data'
sheet_new_book.cell(row=1,column=5).value = 'second_table_data'

# блок фильтрации информации
# check_1 первый фильтр, проверяет, заполнена ли ячейка в двух таблицах
out_date_structure = {} # структура данных содержащая данные для записи в таблицу(данные поделенные по столбцам)
for sell, sell_content in new_control_dict.items(): # sell_content содержит внутри еще 2 словаря, соответствующих двум входящим таблицам
    # sheet_new_book.cell(row=number_row,column=1).value = sell
    condition_list_1 = [] # руда, которая полностью совпадает в 2-х таблицах
    condition_list_2 = [] # руда, которая совпадает по названию, но не совпадает по весу
    condition_list_3 = [] # руда, которая встречается только в первой таблице
    condition_list_4 = [] # руда, которая встречается только во второй таблице
    if len(sell_content) == 2:
        print(sell, 'заполнена в 2-х таблицах')
        list_rock = check_2(sell_content)
        name_set_1 = list_rock['name_1_rock_set'] # множество с рудой которая есть только в первой таблице(итерируемой ячейки)
        name_set_2 = list_rock['name_2_rock_set'] # множество с рудой которая есть только во второй таблице(итерируемой ячейки)
        name_general = list_rock['general_set'] # множество с рудой, которая есть в первой и второй таблице одновременно(итерируемой ячейки)
        if len(name_general) != 0: # если условие верно то нужно проверить, сходится ли руда по весу
            dict_rock_general = check_3(sell_content, name_general)
            matching_data = dict_rock_general['amount_general'] # список содержит данные которые полностью сходятся в 2-х таблицах по текущей ячейке
            dif_list_r = dict_rock_general['difference_rock_list']
            if len(matching_data) != 0:
                for name_rock in matching_data:
                    condition_list_1.append(name_rock)
            if len(dif_list_r) != 0:
                for name_rock in dif_list_r:
                    condition_list_2.append(name_rock)
        if len(name_set_1) != 0:
            for name_rock in name_set_1:
                condition_list_3.append(name_rock)
        if len(name_set_2) != 0:
            for name_rock in name_set_2:
                condition_list_4.append(name_rock)
    else:
        name_tab = list(sell_content.keys())[0]
        rock_dict_tab = sell_content[name_tab]
        name_rock_list_single = []
        for rock_name in rock_dict_tab.keys():
            name_rock_list_single.append(rock_name)
        if name_tab == 'stor_loc_1':
            for name in name_rock_list_single:
                condition_list_3.append(name)
        else:
            for name in name_rock_list_single:
                condition_list_4.append(name)
        print(sell, 'заполнена в таблице', name_tab)
    date_sell = {'condition_list_1': condition_list_1, 'condition_list_2': condition_list_2, 'condition_list_3': condition_list_3, 'condition_list_4': condition_list_4}
    out_date_structure[sell] = date_sell    

# блок записи в excel
number_string = 2        
for key, value in out_date_structure.items():
    matrix = data_to_write(key, value)
    for string in matrix: # записываем данные по ячейки в ексель
        column_write = 1
        for name in string: # записываем данные списка в строку эксель
            if name == '0': #  условие убирает делает клетки с 0 пустыми
                column_write += 1
                continue
            else:
                sheet_new_book.cell(row=number_string,column=column_write).value = name
                column_write += 1
        number_string += 1


new_book.save('output_test_1.xlsx')
new_book.close
print('Запись завершена')

