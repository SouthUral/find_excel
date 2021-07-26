from openpyxl import Workbook
from openpyxl import load_workbook



storage_locations = {}

wb = load_workbook('test_1.xlsx') # load_workbook загружает файл ексель из рабочей директории
# print(wb.sheetnames) # возвращает список всех листов в рабочей книге
ws = wb['Лист1'] # после получения рабочей книги можно обращаться к листам книги по их индексу(названию)
# print(type(ws))
# доступ к ячейкам можно получить напрямую как к ключам рабочего листа
# for row in ws.values:
#     for i in row:
#         print(i)
for row in ws.iter_rows(1, 2,  ):
    print(row)








